import os
import zipfile
import rarfile
from datetime import datetime
import pandas as pd
from pathlib import Path
from docx import Document
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import xlrd
import tempfile
from database import setup_database, save_to_database
from config import PARSER_CONFIG

# Настройка обработки RAR архивов
rarfile.UNRAR_TOOL = "unrar"  # Убедитесь, что unrar установлен в системе

def get_file_type(file_path):
    """Определяет тип файла по расширению"""
    return Path(file_path).suffix.lower()

def extract_text_from_txt(file_path):
    """Извлекает текст из TXT файла"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        print(f"Ошибка чтения TXT файла {file_path}: {str(e)}")
        return ""

def extract_text_from_docx(file_path):
    """Извлекает текст из DOCX файла с обработкой ошибок импорта"""
    try:
        from docx import Document
        try:
            doc = Document(file_path)
            return "\n".join([para.text for para in doc.paragraphs])
        except Exception as e:
            print(f"Ошибка чтения DOCX файла {file_path}: {str(e)}")
            return ""
    except ImportError:
        print(f"Для обработки DOCX файлов требуется установить python-docx: pip install python-docx")
        return ""

def extract_text_from_pdf(file_path):
    """Извлекает текст из PDF файла"""
    try:
        text = ""
        with open(file_path, 'rb') as f:
            reader = PdfReader(f)
            for page in reader.pages:
                text += page.extract_text() + "\n"
        return text.strip()
    except Exception as e:
        print(f"Ошибка чтения PDF файла {file_path}: {str(e)}")
        return ""

def extract_text_from_xlsx(file_path):
    """Извлекает текст из XLSX файла"""
    try:
        text = []
        wb = load_workbook(file_path)
        for sheet in wb:
            for row in sheet.iter_rows(values_only=True):
                text.append(" ".join(str(cell) for cell in row if cell))
        return "\n".join(text)
    except Exception as e:
        print(f"Ошибка чтения XLSX файла {file_path}: {str(e)}")
        return ""

def extract_text_from_xls(file_path):
    """Извлекает текст из XLS файла"""
    try:
        text = []
        wb = xlrd.open_workbook(file_path)
        for sheet in wb.sheets():
            for row in range(sheet.nrows):
                text.append(" ".join(str(sheet.cell(row, col).value) 
                          for col in range(sheet.ncols)))
        return "\n".join(text)
    except Exception as e:
        print(f"Ошибка чтения XLS файла {file_path}: {str(e)}")
        return ""

def process_archive(file_path):
    """Обрабатывает архив и возвращает содержимое вложенных файлов"""
    contents = []
    archive_type = get_file_type(file_path)
    
    with tempfile.TemporaryDirectory() as temp_dir:
        try:
            if archive_type == '.zip':
                with zipfile.ZipFile(file_path, 'r') as z:
                    z.extractall(temp_dir)
                    extracted_files = z.namelist()
                    
            # elif archive_type == '.rar':
            #     with rarfile.RarFile(file_path, 'r') as r:
            #         r.extractall(temp_dir)
            #         extracted_files = r.namelist()
            
            else:
                return contents
            
            # Обрабатываем все извлеченные файлы
            for name in extracted_files:
                extracted_path = os.path.join(temp_dir, name)
                if os.path.isfile(extracted_path):
                    contents += process_file(extracted_path)
        
        except Exception as e:
            print(f"Ошибка обработки архива {file_path}: {str(e)}")
    
    return contents

def process_file(file_path):
    """Обрабатывает файл и возвращает данные для БД"""
    results = []
    file_type = get_file_type(file_path)
    
    # Обработка архивов
    if file_type in ('.zip', '.rar'):
        return process_archive(file_path)
    
    # Обработка поддерживаемых файлов
    if file_type not in PARSER_CONFIG['supported_formats']:
        print(f"Формат {file_type} не поддерживается для файла {file_path}")
        return results
    
    try:
        # Извлечение содержимого
        content = ""
        if file_type == '.txt':
            content = extract_text_from_txt(file_path)
        elif file_type == '.docx':
            content = extract_text_from_docx(file_path)
        elif file_type == '.pdf':
            content = extract_text_from_pdf(file_path)
        elif file_type == '.xlsx':
            content = extract_text_from_xlsx(file_path)
        elif file_type == '.xls':
            content = extract_text_from_xls(file_path)
        
        # Получение метаданных
        stat = os.stat(file_path)
        results.append({
            'path': file_path,
            'type': file_type,
            'content': content,
            'size_kb': round(stat.st_size / 1024, 2),
            'modified': datetime.fromtimestamp(stat.st_mtime),
            'content_length': len(content)
        })
    
    except Exception as e:
        print(f"Ошибка обработки файла {file_path}: {str(e)}")
    
    return results

def scan_directory(root_dir):
    """Рекурсивно сканирует директорию и возвращает результаты"""
    results = []
    
    for dirpath, _, filenames in os.walk(root_dir):
        for filename in filenames:
            file_path = os.path.join(dirpath, filename)
            results += process_file(file_path)
    
    return results

def save_to_csv(data, output_file="output1.csv"):
    """Сохраняет результаты в CSV"""
    if not data:
        print("Нет данных для сохранения!")
        return
    
    try:
        df = pd.DataFrame(data)
        df.to_csv(output_file, index=False, encoding='utf-8-sig')
        print(f"Данные сохранены в {output_file}")
    except Exception as e:
        print(f"Ошибка при сохранении CSV: {str(e)}")

from pathlib import Path

def main():
    setup_database()
    import argparse
    # Определяем путь к папке documents в корне проекта
    project_root = Path(__file__).parent  # Папка проекта, на уровень выше текущего скрипта

    documents_dir = project_root / 'documents'  # Путь к папке documents
    
    extracted_data = scan_directory(documents_dir)
    save_to_csv(extracted_data, PARSER_CONFIG['output_csv'])
    save_to_database(extracted_data)
  
    
    print(f"Обработано {len(extracted_data)} документов")

if __name__ == "__main__":
    main()